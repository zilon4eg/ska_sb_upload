import os
from openpyxl import load_workbook
import shutil


def get_file_in_dir(path):
    # we shall store all the file names in this list
    filelist = []
    for root, dirs, files in os.walk(path):
        for file in files:
            # append the file name to the list
            filelist.append(os.path.join(root, file))

    return filelist


def get_file_in_xlsx(path):
    wb = load_workbook(path)
    ws = wb.active
    filelist = []
    rows = ws.max_row

    for string in range(1, rows + 1):
        filelist.append([ws[f'A{string}'].value, ws[f'C{string}'].value])
    return filelist


if __name__ == '__main__':
    dir_path = r'\\fs\SHARE\Documents\PROJECT-SKA'
    file_path = r'C:\Users\suhorukov.iv\Desktop\Документы ИБ.xlsx'

    files_in_dir = get_file_in_dir(dir_path)
    files_in_xlsx = get_file_in_xlsx(file_path)

    for file_x in files_in_xlsx:
        for file_d in files_in_dir:
            file_x_name = file_x[1]
            file_d_name = file_d[file_d.rfind('\\') + 1:file_d.rfind('.')]

            if str(file_x_name) in str(file_d_name):
                file_name = file_d[file_d.rfind("\\") + 1:]
                # print(file_d, file_name)
                shutil.copyfile(file_d, f'C:\\Users\\suhorukov.iv\\Desktop\\test\\{file_x[0]}_{file_name}')
                print(f'D:\\{file_x[0]}_{file_name}')

